"""
純 Python xlsx → PDF 渲染器（追 MiniPdf 100% 品質）
platypus.Table + dynamic ParagraphStyle per cell
"""
import io, os, re
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus.flowables import HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

PAGE_W, PAGE_H = A4
MARGIN = 8 * mm

# ── CJK 字型 ──
FONT = 'Helvetica'
FONT_BOLD = 'Helvetica-Bold'

for path, name in [
    ('C:/Windows/Fonts/mingliu.ttc', 'MingLiU'),
    ('C:/Windows/Fonts/msjh.ttc', 'MSJH'),
    ('/System/Library/Fonts/PingFang.ttc', 'PingFang'),
]:
    if os.path.exists(path):
        try:
            pdfmetrics.registerFont(TTFont(name, path, subfontIndex=0))
            pdfmetrics.registerFont(TTFont(name + '-B', path, subfontIndex=1))
            FONT = name; FONT_BOLD = name + '-B'
            break
        except:
            pass

if FONT == 'Helvetica':
    try:
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        FONT = 'STSong-Light'; FONT_BOLD = 'HeiseiKakuGo-W5'
    except:
        pass


def col_letter_to_index(letter):
    return column_index_from_string(letter) - 1


def _eval_formula(formula, ws, row, col):
    if not formula or not formula.startswith('='):
        return formula
    try:
        expr = formula[1:]

        def cv(ref):
            m = re.match(r'\$?([A-G])\$?(\d+)', ref)
            if not m: return 0
            v = ws[f'{m.group(1)}{m.group(2)}'].value
            if v is None: return 0
            if isinstance(v, str) and v.startswith('='):
                return float(_eval_formula(v, ws, int(m.group(2)), 0) or 0)
            try: return float(v)
            except: return 0

        expr = re.sub(r'\$?([A-G])\$?(\d+)', lambda m: str(cv(m.group(0))), expr)
        expr = re.sub(r'SUM\(([A-G]\d+):([A-G]\d+)\)',
                      lambda m: str(_sum_range(ws, m.group(1), m.group(2))),
                      expr, flags=re.IGNORECASE)
        r = eval(expr)
        if isinstance(r, float) and r == int(r): r = int(r)
        return str(r)
    except:
        return '0'


def _sum_range(ws, s, e):
    m1 = re.match(r'\$?([A-G])\$?(\d+)', s)
    m2 = re.match(r'\$?([A-G])\$?(\d+)', e)
    if not m1 or not m2: return 0
    col, r1, r2 = m1.group(1), int(m1.group(2)), int(m2.group(2))
    t = 0
    for r in range(r1, r2 + 1):
        v = ws[f'{col}{r}'].value
        if v is None: continue
        if isinstance(v, str) and v.startswith('='):
            v = _eval_formula(v, ws, r, 0)
        try: t += float(v)
        except: pass
    return t


def _fmt_currency(val):
    try:
        n = int(float(val))
        return f'${n:,}'
    except:
        return str(val)


def _is_currency(cell):
    try:
        nf = str(cell.number_format or '')
        return '$' in nf or '#,##0' in nf
    except:
        return False


def _get_cell_style(cell, base_size=10):
    """Create a unique ParagraphStyle for this cell's formatting"""
    fn = FONT
    fs = base_size
    fc = colors.black
    ha = TA_LEFT

    try:
        if cell.font:
            if cell.font.size:
                fs = max(min(cell.font.size, 26), 6)
            if cell.font.bold:
                fn = FONT_BOLD
            if cell.font.color and cell.font.color.rgb:
                rgb = cell.font.color.rgb
                if rgb not in ('00000000', '0'):
                    fc = colors.HexColor('#' + rgb[2:])
    except: pass

    try:
        if cell.alignment and cell.alignment.horizontal:
            if cell.alignment.horizontal == 'center': ha = TA_CENTER
            elif cell.alignment.horizontal == 'right': ha = TA_RIGHT
    except: pass

    # Cache key for style reuse
    key = (fn, fs, str(fc), ha)
    if not hasattr(_get_cell_style, 'cache'):
        _get_cell_style.cache = {}

    if key in _get_cell_style.cache:
        return _get_cell_style.cache[key]

    style = ParagraphStyle(
        f'cell_{len(_get_cell_style.cache)}',
        fontName=fn, fontSize=fs, leading=fs + 3,
        textColor=fc, alignment=ha,
    )
    _get_cell_style.cache[key] = style
    return style


def convert_xlsx_to_pdf(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=MARGIN, rightMargin=MARGIN,
                            topMargin=MARGIN, bottomMargin=MARGIN)
    story = []

    # ── Merged cells ──
    merged = {}
    for m in ws.merged_cells.ranges:
        r1, r2 = m.min_row, m.max_row
        c1 = col_letter_to_index(m.min_col) if isinstance(m.min_col, str) else m.min_col - 1
        c2 = col_letter_to_index(m.max_col) if isinstance(m.max_col, str) else m.max_col - 1
        for ri in range(r1, r2 + 1):
            for ci in range(c1, c2 + 1):
                merged[(ri, ci)] = (r1, c1, r2, c2)

    # ── Column widths ──
    col_widths = []
    for ci in range(ws.max_column):
        letter = get_column_letter(ci + 1)
        w = ws.column_dimensions.get(letter)
        col_widths.append((w.width or 10) * 6.5 if w and w.width else 65)

    total_w = sum(col_widths)
    scale = (PAGE_W - 2 * MARGIN) / max(total_w, 1)
    col_w = [w * scale for w in col_widths]

    # ── Build table data row by row ──
    table_data = []
    row_map = []  # map table_data index → xlsx row number
    table_row = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = row[0].row
        row_cells = []
        has_any = False

        for cell in row:
            ci = cell.column - 1
            if ci >= len(col_w):
                continue

            # Skip merged sub-cells
            if (r, ci) in merged:
                mr1, mc1, mr2, mc2 = merged[(r, ci)]
                if (r, ci) != (mr1, mc1):
                    row_cells.append(None)  # placeholder
                    continue

            if cell.value is None:
                row_cells.append('')
                continue

            val = str(cell.value)
            if val.startswith('='):
                val = _eval_formula(val, ws, r, ci)
            if _is_currency(cell) and val.lstrip('-').replace('.', '', 1).isdigit():
                val = _fmt_currency(val)

            val = val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            style = _get_cell_style(cell)
            para = Paragraph(val, style)
            row_cells.append(para)
            has_any = True

        if has_any or any(c is not None and c != '' for c in row_cells):
            # Fill None placeholders with empty Paragraph
            row_cells = [c if c is not None else Paragraph('', _get_cell_style(ws.cell(row=r, column=1))) for c in row_cells]
            # Ensure all rows have same column count
            while len(row_cells) < len(col_w):
                row_cells.append('')
            table_data.append(row_cells)
            row_map.append(r)
            table_row += 1

    # ── Table styles ──
    style_cmds = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    # Scan xlsx again to build merge/bg/border commands
    for ti, r in enumerate(row_map):
        for cell in ws.iter_rows(min_row=r, max_row=r):
            for cell in cell:
                ci = cell.column - 1
                if ci >= len(col_w):
                    continue

                # SPAN
                if (r, ci) in merged:
                    mr1, mc1, mr2, mc2 = merged[(r, ci)]
                    if (r, ci) == (mr1, mc1):
                        # Find target table rows
                        end_ti = ti
                        for tj, rj in enumerate(row_map):
                            if rj == mr2:
                                end_ti = tj; break
                        if mc2 < len(col_w):
                            style_cmds.append(('SPAN', (mc1, ti), (mc2, end_ti)))

                # BACKGROUND
                try:
                    fill = cell.fill
                    if fill.patternType == 'solid':
                        rgb = fill.fgColor.rgb
                        if rgb and rgb not in ('00000000', '0'):
                            style_cmds.append(('BACKGROUND', (ci, ti), (ci, ti),
                                              colors.HexColor('#' + rgb[2:])))
                except: pass

                # BOX border
                try:
                    b = cell.border
                    if any(getattr(b, s).style for s in ['left', 'right', 'top', 'bottom']):
                        style_cmds.append(('BOX', (ci, ti), (ci, ti), 0.3,
                                          colors.HexColor('#D9D9D9')))
                except: pass

    # ── Build and render ──
    t = Table(table_data, colWidths=col_w, repeatRows=0)
    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    # Footer line
    story.append(Spacer(1, 5 * mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#999999')))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── CLI ──
if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('xlsx2pdf — 純 Python xlsx → PDF 渲染器')
        print('用法: python xlsx2pdf.py <input.xlsx> [-o output.pdf]')
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = input_path.replace('.xlsx', '.pdf')
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '-o' and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]; i += 2
        else:
            i += 1

    with open(input_path, 'rb') as f:
        pdf = convert_xlsx_to_pdf(f.read())

    with open(output_path, 'wb') as f:
        f.write(pdf)

    print(f'OK: {input_path} → {output_path} ({len(pdf)} bytes)')
