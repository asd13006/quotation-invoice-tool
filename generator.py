"""
裝修報價單 Excel 生成器 — 跟目標格式
"""
from collections import OrderedDict
from openpyxl.utils import get_column_letter
from styles import *

CN_NUMS = ['一', '二', '三', '四', '五', '六', '七']


def generate_quotation(ws, data, title='報價單'):
    ws.title = '裝修發票' if title == '發票' else SHEET_NAME
    set_col_widths(ws, COL_WIDTHS)

    # 按 category 歸入 section
    section_items = {sec['num']: [] for sec in SECTIONS}
    for it in data.get('items', []):
        cat = it.get('category', '')
        for sec in SECTIONS:
            if cat == sec['cat']:
                section_items[sec['num']].append(it)
                break
        else:
            section_items[7].append(it)

    r = 1

    # ═══ Row 1: 標題 ═══
    title_border = Border(bottom=Side(style='thin'))
    # 先对所有格加底線，再合併
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        apply_cell(ws, f'{col}{r}', border=title_border)
    ws.merge_cells(f'A{r}:G{r}')
    apply_cell(ws, f'A{r}', value=title, font=FONT_TITLE,
               alignment=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[r].height = 40.0
    r = 3  # skip row 2

    # ═══ Rows 3-6: 資訊 ═══
    info_left = [
        ('工程名稱：', 'project_name'),
        ('客戶姓名：', 'owner_name'),
        ('工程地址：', 'address'),
        ('裝修公司：', 'company_name'),
    ]
    info_right = [
        ('報價單號：', 'quotation_no'),
        ('報價日期：', 'date'),
        ('有效期：', 'validity'),
        ('版本：', 'version'),
    ]
    for i in range(4):
        row = r + i
        apply_cell(ws, f'A{row}', value=info_left[i][0], font=FONT_INFO_LABEL)
        apply_cell(ws, f'B{row}', value=data.get(info_left[i][1], '') or '-', font=FONT_INFO_VALUE)
        ws.merge_cells(f'B{row}:D{row}')
        apply_cell(ws, f'E{row}', value=info_right[i][0], font=FONT_INFO_LABEL)
        apply_cell(ws, f'F{row}', value=data.get(info_right[i][1], '') or '-', font=FONT_INFO_VALUE)
        ws.merge_cells(f'F{row}:G{row}')
        ws.row_dimensions[row].height = ROW_HEIGHT_INFO
    r += 4

    # ═══ Row 8: 全頁唯一表頭 ═══
    r += 1  # spacer row 7
    _write_table_header(ws, r)
    r += 1

    # ═══ Sections ═══
    subtotal_cells = []
    section_counter = 0

    for sec in SECTIONS:
        items = section_items[sec['num']]
        if not items:
            continue

        section_title = f'{CN_NUMS[section_counter]}、 {sec["title"]}'
        section_counter += 1

        # Section header (merged A:G, 冇框, 淺灰底)
        ws.merge_cells(f'A{r}:G{r}')
        apply_cell(ws, f'A{r}', value=section_title, font=FONT_SECTION,
                   fill=FILL_SECTION, alignment=ALIGN_LEFT)
        ws.row_dimensions[r].height = 20.5
        r += 1

        # Items
        item_start = r
        for i, it in enumerate(items):
            seq = f'{section_counter}.{i + 1}'
            apply_cell(ws, f'A{r}', value=seq, font=FONT_ITEM, alignment=ALIGN_CENTER, border=BORDER_GRAY)
            apply_cell(ws, f'B{r}', value=it['description'], font=FONT_ITEM, alignment=ALIGN_LEFT, border=BORDER_GRAY)
            apply_cell(ws, f'C{r}', value=it.get('quantity', 1), font=FONT_ITEM, alignment=ALIGN_CENTER, border=BORDER_GRAY)
            apply_cell(ws, f'D{r}', value=it.get('unit', '項'), font=FONT_ITEM, alignment=ALIGN_CENTER, border=BORDER_GRAY)
            apply_cell(ws, f'E{r}', value=it.get('unit_price', 0), font=FONT_ITEM, alignment=ALIGN_RIGHT, border=BORDER_GRAY, number_format=FMT_CURRENCY)
            apply_cell(ws, f'F{r}', value=f'=C{r}*E{r}', font=FONT_ITEM, alignment=ALIGN_RIGHT, border=BORDER_GRAY, number_format=FMT_CURRENCY)
            apply_cell(ws, f'G{r}', value=it.get('remark', '') or '-', font=FONT_ITEM, alignment=ALIGN_LEFT, border=BORDER_GRAY)
            ws.row_dimensions[r].height = ROW_HEIGHT_ITEM
            r += 1
        item_end = r - 1

        # Subtotal: "小計：" in E, formula in F — thin 全包
        apply_cell(ws, f'E{r}', value='小計：', font=FONT_SUBTOTAL,
                   fill=FILL_SECTION, alignment=ALIGN_RIGHT)
        apply_cell(ws, f'F{r}', value=f'=SUM(F{item_start}:F{item_end})',
                   font=FONT_SUBTOTAL, fill=FILL_SECTION, alignment=ALIGN_RIGHT,
                   number_format=FMT_CURRENCY)
        ws.row_dimensions[r].height = ROW_HEIGHT_ITEM
        subtotal_cells.append(f'F{r}')
        r += 2  # spacer

    # ═══ 總計 ═══
    ws.merge_cells(f'A{r}:E{r}')
    apply_cell(ws, f'A{r}', value='總工程預算總計 (HKD)：',
               font=FONT_TOTAL, fill=FILL_SECTION, alignment=ALIGN_RIGHT)
    apply_cell(ws, f'F{r}', value='=' + '+'.join(subtotal_cells),
               font=FONT_TOTAL, fill=FILL_SECTION, alignment=ALIGN_RIGHT, number_format=FMT_CURRENCY)
    ws.row_dimensions[r].height = ROW_HEIGHT_TOTAL
    grand_total_row = r
    r += 1

    # 訂金 + 尾款（只限發票，且訂金 > 0）
    deposit = data.get('deposit', 0)
    if title == '發票' and deposit > 0:
        apply_cell(ws, f'A{r}', value='訂金 (Deposit)：', font=FONT_DEPOSIT_LABEL, alignment=ALIGN_RIGHT)
        ws.merge_cells(f'A{r}:E{r}')
        apply_cell(ws, f'F{r}', value=deposit, font=FONT_DEPOSIT_VALUE,
                   alignment=ALIGN_RIGHT, number_format=FMT_CURRENCY)
        ws.row_dimensions[r].height = ROW_HEIGHT_ITEM
        deposit_row = r
        r += 1

        apply_cell(ws, f'A{r}', value='應付尾款 (Balance Due)：', font=FONT_DEPOSIT_LABEL, alignment=ALIGN_RIGHT)
        ws.merge_cells(f'A{r}:E{r}')
        apply_cell(ws, f'F{r}', value=f'=F{grand_total_row}-F{deposit_row}',
                   font=FONT_BALANCE, alignment=ALIGN_RIGHT, number_format=FMT_CURRENCY)
        ws.row_dimensions[r].height = ROW_HEIGHT_ITEM
        r += 1

    r += 1

    # ═══ 付款 ═══
    show_payment = data.get('show_payment', True)
    show_terms = data.get('show_terms', True)

    if show_payment:
        apply_cell(ws, f'A{r}', value='工程付款階段說明：', font=FONT_TERMS_LABEL)
        r += 1

        # Payment header
        for col, hdr in [('A', '付款期數'), ('B', '比例'), ('C', '金額 (HKD)'), ('D', '付款條件說明')]:
            apply_cell(ws, f'{col}{r}', value=hdr, font=FONT_PAYMENT_HEADER, border=BORDER_GRAY)
        ws.merge_cells(f'D{r}:G{r}')
        ws.row_dimensions[r].height = ROW_HEIGHT_PAYMENT
        r += 1

        payments = data.get('payments', [
            {'label': '第一期 (簽約訂金)', 'pct': 0.2, 'desc': '確認施工圖則、清拆及工程保護進場前付清'},
            {'label': '第二期 (泥水水電)', 'pct': 0.4, 'desc': '水電管線暗埋及泥水工程完工、驗收後付清'},
            {'label': '第三期 (木工油漆)', 'pct': 0.3, 'desc': '木工現場結構完成、油漆進場及訂造傢俬進場前付清'},
            {'label': '第四期 (工程尾數)', 'pct': 0.1, 'desc': '全屋工程完工、清潔交吉及驗收完成後付清'},
        ])

        for p in payments:
            apply_cell(ws, f'A{r}', value=p['label'], font=FONT_PAYMENT, border=BORDER_GRAY)
            apply_cell(ws, f'B{r}', value=p.get('label_pct', f'總金額之 {int(p["pct"]*100)}%'),
                       font=FONT_PAYMENT, alignment=ALIGN_CENTER, border=BORDER_GRAY)
            apply_cell(ws, f'C{r}', value=f'=F{grand_total_row}*{p["pct"]}',
                       font=FONT_PAYMENT_AMT, alignment=ALIGN_RIGHT, number_format=FMT_CURRENCY, border=BORDER_GRAY)
            apply_cell(ws, f'D{r}', value=p['desc'], font=FONT_PAYMENT, border=BORDER_GRAY)
            ws.merge_cells(f'D{r}:G{r}')
            ws.row_dimensions[r].height = ROW_HEIGHT_PAYMENT
            r += 1
        r += 1

    # ═══ 條款 ═══
    if show_terms:
        apply_cell(ws, f'A{r}', value='備註及條款說明：', font=FONT_TERMS_LABEL)
        r += 1

        terms = data.get('terms', [
            '本報價單所列工程，於完工後均享有 1 年結構及防漏水工程保養期。',
            '報價已包含工程期間之第三者責任保險及僱員補償保險（勞保）。',
            '所有加減工程（Variation Order）必須經雙方書面確認金額及工期變更後方可施工，口頭協議一律無效。',
            '電位計算定義：單位插座計1個位，雙位插座計1.5個位，燈制與燈位分開計算，現場放線時由雙方核對確認。',
            '預計工期為 90 個工作天（不計星期日及公眾假期），若因公司原因無故延誤，每逾期一日補償業主 HKD $500。',
        ])

        for i, t in enumerate(terms, 1):
            apply_cell(ws, f'A{r}', value=f'{i}. {t}', font=FONT_TERMS, alignment=ALIGN_WRAP)
            ws.merge_cells(f'A{r}:G{r}')
            ws.row_dimensions[r].height = ROW_HEIGHT_TERMS
            r += 1

    setup_a4_print(ws)

def _write_table_header(ws, row):
    headers = ['項目編號', '工程項目及說明', '數量', '單位', '單價 (HKD)', '複價 (HKD)', '備註']
    alignments = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_CENTER,
                  ALIGN_RIGHT, ALIGN_RIGHT, ALIGN_LEFT]
    for i, (hdr, al) in enumerate(zip(headers, alignments)):
        col = get_column_letter(i + 1)
        apply_cell(ws, f'{col}{row}', value=hdr, font=FONT_TABLE_HEADER,
                   alignment=al, border=BORDER_HEADER_BOTTOM)
    ws.row_dimensions[row].height = 22.0
