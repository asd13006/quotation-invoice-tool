"""
模板格式定義 — Interior Design Quotation Template
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ── 字型 (全用微軟正黑體) ──────────────────────────
FONT_TITLE = Font(name='Microsoft JhengHei', size=26, bold=True, color='1a1a1a')
FONT_INFO_LABEL = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_INFO_VALUE = Font(name='Microsoft JhengHei', size=10, bold=False, color='333333')
FONT_SECTION = Font(name='Microsoft JhengHei', size=11, bold=True, color='1a1a1a')
FONT_TABLE_HEADER = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_ITEM = Font(name='Microsoft JhengHei', size=10, bold=False, color='333333')
FONT_SUBTOTAL = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_TOTAL = Font(name='Microsoft JhengHei', size=11, bold=True, color='1a1a1a')
FONT_PAYMENT_HEADER = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_PAYMENT = Font(name='Microsoft JhengHei', size=10, bold=False, color='333333')
FONT_PAYMENT_AMT = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_TERMS_LABEL = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_TERMS = Font(name='Microsoft JhengHei', size=10, bold=False, color='555555')
FONT_DEPOSIT_LABEL = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')
FONT_DEPOSIT_VALUE = Font(name='Microsoft JhengHei', size=10, bold=False, color='333333')
FONT_BALANCE = Font(name='Microsoft JhengHei', size=10, bold=True, color='1a1a1a')

# ── 填滿顏色 ──────────────────────────────────────────
# 列印友善：只用淺灰做層次，極低墨水成本
FILL_SECTION = PatternFill(start_color='00F2F2F2', end_color='00F2F2F2', fill_type='solid')
FILL_TABLE_HEADER = PatternFill(start_color='00F2F2F2', end_color='00F2F2F2', fill_type='solid')
FILL_ROW_ALT = PatternFill(fill_type=None)
FILL_HEADER_BG = PatternFill(fill_type=None)
FILL_TOTAL = PatternFill(fill_type=None)
FILL_NONE = PatternFill(fill_type=None)

# ── 對齊方式 ──────────────────────────────────────────
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── 邊框 ──────────────────────────────────────────────
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
# 項目用淺灰框
BORDER_GRAY = Border(
    left=Side(style='thin', color='FFF2F2F2'),
    right=Side(style='thin', color='FFF2F2F2'),
    top=Side(style='thin', color='FFF2F2F2'),
    bottom=Side(style='thin', color='FFF2F2F2'),
)
BORDER_TOTAL = Border(
    top=Side(style='thin'),
    bottom=Side(style='double'),
)
# Section 標題：L+R+T thin, B none
BORDER_SECTION = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
)
# 表頭：only B thin
BORDER_HEADER_BOTTOM = Border(
    bottom=Side(style='thin'),
)
# 小計：E column has T thin, F no border
BORDER_SUBTOTAL_E = Border(
    top=Side(style='thin'),
)
BORDER_NONE = Border()

# ── 數字格式 ──────────────────────────────────────────
FMT_CURRENCY = '$#,##0'
FMT_PERCENT = '0%'

# ── 行高 ──────────────────────────────────────────────
ROW_HEIGHT_TITLE = 40.0
ROW_HEIGHT_INFO = 22.0
ROW_HEIGHT_SECTION = 28.0
ROW_HEIGHT_TABLE_HEADER = 25.0
ROW_HEIGHT_ITEM = 22.0
ROW_HEIGHT_TOTAL = 30.0
ROW_HEIGHT_PAYMENT = 22.0
ROW_HEIGHT_TERMS = 22.0

# ── 欄寬 ──────────────────────────────────────────────
COL_WIDTHS = {
    'A': 12.0, 'B': 40.0, 'C': 7.0, 'D': 6.0,
    'E': 12.0, 'F': 14.0, 'G': 24.0,
}

# ── 7 大區段定義 ──────────────────────────────────────
SECTIONS = [
    {'num': 1, 'title': '前期及清拆工程', 'cat': '清拆工程'},
    {'num': 2, 'title': '水電工程', 'cat': '水電工程'},
    {'num': 3, 'title': '泥水工程', 'cat': '泥水工程'},
    {'num': 4, 'title': '油漆工程', 'cat': '油漆工程'},
    {'num': 5, 'title': '木工工程', 'cat': '木工工程'},
    {'num': 6, 'title': '安裝代工', 'cat': '安裝代工'},
    {'num': 7, 'title': '雜項', 'cat': '雜項'},
]

SHEET_NAME = '裝修報價單'

# ── Helper Functions ──────────────────────────────────

def apply_cell(ws, coord, value=None, font=None, fill=None, alignment=None,
               border=None, number_format=None):
    """Helper: 設定儲存格值 + 格式"""
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def set_col_widths(ws, widths_dict):
    """Helper: 設定欄寬"""
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def setup_a4_print(ws):
    """Helper: 設定 A4 滿版列印"""
    # A4 paper
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    # Fit to 1 page wide x 1 page tall
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToPage = True
    # Margins in inches
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    # Print area
    ws.print_area = f'A1:G{ws.max_row}'
    ws.sheet_properties.pageSetUpPr = None
