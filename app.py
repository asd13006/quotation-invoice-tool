"""裝修報價單/發票助手 — Flask v3.4.0 (+dashboard, +drive save, +whatsapp, +project tracking)"""
import io, os, re, uuid, random, smtplib, time, json as _json
from email.mime.text import MIMEText
from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from generator import generate_quotation
from drive_sync import list_projects, get_project_data, upload_file
from projects import register_project, list_projects_local, get_dashboard_stats, update_status, toggle_payment, get_project, sync_from_drive, parse_drive_files

app = Flask(__name__)

# ── Cache control: prevent stale assets ──
@app.after_request
def add_cache_headers(response):
    if request.path.startswith('/static/'):
        # Cache static assets for 1 hour, but allow revalidation
        response.headers['Cache-Control'] = 'public, max-age=3600'
    else:
        # Never cache HTML pages
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

with open(os.path.join(os.path.dirname(__file__), 'VERSION')) as f:
    _VERSION = f.read().strip()

_preview_cache = {}


# ── Static files ──
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)


@app.route('/manifest.json')
def manifest():
    return send_file('static/manifest.json', mimetype='application/manifest+json')


@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')


# ═══════════════ ROUTES ═══════════════

@app.route('/')
def index():
    if not _check_auth():
        return '<script>location.href="/login"</script>'
    for spa_path in [
        os.path.join(os.path.dirname(__file__), 'static', 'dist-spa', 'index.html'),
        os.path.join(os.path.dirname(__file__), 'frontend', 'index.html'),
    ]:
        if os.path.exists(spa_path):
            return open(spa_path, encoding='utf-8').read()
    return render_template('index.html', version=_VERSION)


@app.route('/dashboard')
def dashboard_page():
    if not _check_auth():
        return '<script>location.href="/login"</script>'
    return render_template('dashboard.html', version=_VERSION)


@app.route('/projects-manage')
def projects_manage_page():
    if not _check_auth():
        return '<script>location.href="/login"</script>'
    return render_template('projects_manage.html', version=_VERSION)


@app.route('/minipdf')
def minipdf_page():
    if not _check_auth():
        return '<script>location.href="/login"</script>'
    return render_template('minipdf.html', version=_VERSION)


@app.route('/minipdf/')
@app.route('/minipdf/<path:filename>')
def serve_minipdf(filename='index.html'):
    return send_from_directory('static/minipdf', filename)


@app.route('/generate', methods=['POST'])
def generate():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({'error': '無效的請求資料'}), 400
    data = body.get('data', {})
    if not data.get('items'):
        return jsonify({'error': '請至少填寫一個工程項目'}), 400
    try:
        gen_type = body.get('type', 'quotation')
        doc_title = '發票' if gen_type == 'invoice' else '報價單'
        wb = Workbook()
        generate_quotation(wb.active, data, title=doc_title)
        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_bytes = xlsx_buf.getvalue()
        pid = uuid.uuid4().hex[:8]
        fname = _make_filename(data, doc_title)

        # Register project for lifecycle tracking
        try:
            proj = register_project(data, gen_type)
            project_id = proj.get('id', '')
        except Exception:
            project_id = ''

        html = _build_download_page(pid, doc_title, data, fname, project_id)
        _preview_cache[pid] = {
            'html': html, 'xlsx': xlsx_bytes,
            '_filename': fname + '.xlsx', '_project_id': project_id,
            '_gen_type': gen_type, '_data': data,
        }
        return jsonify({'preview_id': pid, 'status': 'ok'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/preview/<pid>')
def preview(pid):
    entry = _preview_cache.get(pid)
    return (entry['html'] if entry else '<h1>預覽已過期</h1>')


@app.route('/download/<pid>/excel')
def download_excel(pid):
    entry = _preview_cache.get(pid)
    if not entry or not entry.get('xlsx'):
        return 'Not found', 404
    fname = entry.get('_filename', '報價單.xlsx')
    return send_file(io.BytesIO(entry['xlsx']),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/api/whatsapp/<pid>')
def whatsapp_share(pid):
    """Return WhatsApp share info"""
    entry = _preview_cache.get(pid)
    if not entry:
        return jsonify({'error': 'preview expired'}), 404
    data = entry.get('_data', {})
    gen_type = entry.get('_gen_type', 'quotation')
    doc_title = '發票' if gen_type == 'invoice' else '報價單'
    addr = (data.get('address') or data.get('project_name') or '').strip()
    total = sum((it.get('quantity', 1) or 1) * (it.get('unit_price', 0) or 0) for it in data.get('items', []))
    text = (
        f"{doc_title} — {addr}\n"
        f"單號：{data.get('quotation_no', '-')}\n"
        f"總額：${total:,.0f} HKD\n"
        f"日期：{data.get('date', '-')}"
    )
    return jsonify({
        'text': text,
        'filename': entry.get('_filename', ''),
        'download_url': f'/download/{pid}/excel',
    })


@app.route('/save-drive/<pid>', methods=['POST'])
def save_to_drive(pid):
    """Upload generated Excel to Google Drive"""
    entry = _preview_cache.get(pid)
    if not entry or not entry.get('xlsx'):
        return jsonify({'error': 'No file to save'}), 404
    try:
        fname = entry.get('_filename', 'quotation.xlsx')
        folder_id = request.args.get('folder', '')
        result = upload_file(entry['xlsx'], fname, folder_id if folder_id else None)
        return jsonify({'status': 'ok', 'file_id': result.get('id', ''),
                        'web_link': f"https://drive.google.com/file/d/{result.get('id','')}/view"})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════ UPLOAD ═══════════════

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '請選擇 .xlsx 檔案'}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        result = {'items': [], 'payments': [], 'terms': [], 'deposit': 0,
                  '_filename': file.filename, 'show_payment': False, 'show_terms': False}
        header_map = {
            '工程名稱': 'project_name', '工程地址': 'address',
            '客戶姓名': 'owner_name', '業主姓名': 'owner_name',
            '裝修公司': 'company_name', '報價單號': 'quotation_no',
            '報價日期': 'date', '有效期': 'validity', '版本': 'version',
        }
        for r in range(1, 21):
            for c in range(1, 10):
                label = str(ws.cell(row=r, column=c).value or '').strip()
                if not label:
                    continue
                for kw, key in header_map.items():
                    if kw not in label:
                        continue
                    val = str(ws.cell(row=r, column=c + 1).value or '')
                    if val in ('None', '', '-'):
                        val = str(ws.cell(row=r, column=c + 2).value or '')
                    if val in ('None', '', '-'):
                        for cc in range(c + 1, 10):
                            v = str(ws.cell(row=r, column=cc).value or '')
                            if v not in ('None', '', '-'):
                                val = v
                                break
                    if val and val not in ('None', '', '-'):
                        result[key] = val
                        break
        for r in range(1, 21):
            for c in range(1, 10):
                label = str(ws.cell(row=r, column=c).value or '').strip()
                if not label:
                    continue
                for kw, key in [('工程地址', 'address'), ('報價日期', 'date')]:
                    if kw in label and ('：' in label or ':' in label):
                        sep = '：' if '：' in label else ':'
                        parts = label.split(sep, 1)
                        if len(parts) > 1 and parts[1].strip():
                            result[key] = parts[1].strip()

        items = _parse_items_universal(ws)
        result['items'] = [{
            **it,
            'category': _auto_categorize(it['description']),
            'is_additional': False,
        } for it in items]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if not row or row[0] is None:
                continue
            a = str(ws.cell(row=row[0].row, column=1).value or '')
            r = row[0].row
            if '付款期數' in a:
                pr = r + 1
                while pr <= ws.max_row:
                    pa = str(ws.cell(row=pr, column=1).value or '')
                    pb = str(ws.cell(row=pr, column=2).value or '')
                    pd = str(ws.cell(row=pr, column=4).value or '')
                    if pa and pa != 'None' and '備註' not in pa and '條款' not in pa:
                        pct_m = re.search(r'(\d+)%', pb)
                        pct = int(pct_m.group(1)) if pct_m else 25
                        result['payments'].append({
                            'label': pa, 'pct': pct,
                            'label_pct': pb if pb not in ('None', '') else f'總金額之 {pct}%',
                            'desc': pd if pd not in ('None', '') else '',
                        })
                        pr += 1
                    else:
                        break
                result['show_payment'] = len(result['payments']) > 0
            if '備註及條款' in a:
                tr = r + 1
                while tr <= ws.max_row:
                    ta = str(ws.cell(row=tr, column=1).value or '')
                    if ta and ta != 'None' and '付款' not in ta:
                        ta = re.sub(r'^\d+\.\s*', '', ta).strip()
                        if len(ta) > 3:
                            result['terms'].append(ta)
                        tr += 1
                    else:
                        tr += 1
                        if tr - r > 20:
                            break
                result['show_terms'] = len(result['terms']) > 0
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════ LOGIN ═══════════════

_pending_codes = {}
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')


def _check_auth():
    allowed = os.environ.get('LOGIN_EMAIL', '')
    if not allowed:
        return True
    return request.cookies.get('auth') == allowed


def _send_code(email):
    code = ''.join(random.choices('0123456789', k=6))
    _pending_codes[email] = {'code': code, 'expires': time.time() + 300, 'tries': 0, 'max_tries': 5}
    msg = MIMEText(f'你的登入驗證碼：\n\n{code}\n\n5 分鐘內有效。', 'plain', 'utf-8')
    msg['Subject'] = '工程單助手 — 登入驗證碼'
    msg['From'] = SMTP_USER
    msg['To'] = email
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg)


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    allowed = os.environ.get('LOGIN_EMAIL', '')
    if not allowed:
        return '<script>location.href="/"</script>'
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        code = request.form.get('code', '').strip()
        if not code:
            if email != allowed:
                return _login_page_html(error='此郵箱未獲授權')
            try:
                _send_code(email)
                return _login_page_html(email=email, sent=True)
            except Exception as e:
                return _login_page_html(error=str(e))
        if email == allowed and code:
            pending = _pending_codes.get(email)
            if not pending or time.time() >= pending['expires']:
                return _login_page_html(email=email, sent=True, error='驗證碼已過期')
            pending['tries'] += 1
            if pending['code'] == code:
                del _pending_codes[email]
                resp = app.make_response('<script>location.href="/"</script>')
                resp.set_cookie('auth', allowed, max_age=60 * 60 * 24 * 30)
                return resp
            remaining = pending['max_tries'] - pending['tries']
            if remaining <= 0:
                del _pending_codes[email]
                return _login_page_html(error='驗證碼錯誤次數過多')
            return _login_page_html(email=email, sent=True, error=f'驗證碼錯誤，仲有 {remaining} 次機會')
    return _login_page_html()


def _login_page_html(email='', sent=False, error=''):
    css = '''
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Microsoft JhengHei',sans-serif;background:linear-gradient(135deg,#e0e7f0,#dce3ed,#e8edf5,#d5ddef);display:flex;justify-content:center;align-items:center;min-height:100vh;padding:20px}
.card{background:rgba(255,255,255,0.55);backdrop-filter:blur(24px);border-radius:18px;padding:44px 36px;text-align:center;box-shadow:0 12px 40px rgba(0,0,0,.08);max-width:400px;width:100%}
.icon{font-size:40px;margin-bottom:12px}
h1{font-size:20px;margin-bottom:4px;color:#1e293b;font-weight:700}
h1 span{font-size:12px;color:#94a3b8;display:block;margin-top:4px;font-weight:400}
p.desc{color:#64748b;font-size:14px;margin:16px 0 20px}
.email-display{background:rgba(59,130,246,0.08);color:#3b82f6;padding:6px 14px;border-radius:20px;font-size:13px;display:inline-block;margin-bottom:16px}
input{padding:13px 16px;border:1px solid rgba(0,0,0,.1);border-radius:10px;font-size:16px;width:100%;margin-bottom:12px;background:rgba(255,255,255,.6);color:#1e293b;font-family:'Microsoft JhengHei',sans-serif}
.code-input{font-size:24px;text-align:center;letter-spacing:8px;font-family:'SF Mono',Consolas,monospace}
input:focus{outline:none;border-color:#3b82f6;box-shadow:0 0 0 3px rgba(59,130,246,.15)}
button{padding:13px 28px;border:none;border-radius:10px;background:#3b82f6;color:#fff;font-size:15px;font-weight:600;cursor:pointer;width:100%;box-shadow:0 4px 14px rgba(59,130,246,.25)}
button:hover{background:#2563eb}
.error{color:#ef4444;font-size:13px;margin-bottom:10px}
.success{color:#22c55e;font-size:13px;margin-bottom:10px}
.resend{color:#94a3b8;font-size:12px;margin-top:12px}
.resend a{color:#3b82f6;text-decoration:none}
@media(prefers-color-scheme:dark){
body{background:linear-gradient(135deg,#0f172a,#1a1f35,#0d1525)}
h1{color:#e2e8f0}.desc{color:#94a3b8}
.card{background:rgba(30,41,59,.6)}
input{background:rgba(30,41,59,.7);border-color:rgba(255,255,255,.1);color:#e2e8f0}
}
'''
    error_html = f'<div class="error">{error}</div>' if error else ''
    success_html = '<div class="success">驗證碼已發送到你的郵箱</div>' if sent else ''
    if sent:
        body = f'''
<div class="icon">&#x1F4EC;</div>
<h1>驗證碼已發送<span>請檢查郵箱</span></h1>
<div class="email-display">{email}</div>
{success_html}{error_html}
<form method="POST"><input type="hidden" name="email" value="{email}">
<input class="code-input" type="text" name="code" placeholder="000000" maxlength="6" inputmode="numeric" autocomplete="one-time-code" autofocus>
<button type="submit">驗證登入</button></form>
<p class="resend">收唔到？<a href="/login">重新發送</a></p>'''
    else:
        body = f'''
<div class="icon">&#x1F4CB;</div>
<h1>工程單助手<span>Email 驗證登入</span></h1>
<p class="desc">輸入授權郵箱，接收一次性驗證碼</p>
{error_html}
<form method="POST">
<input type="email" name="email" placeholder="your@email.com" autofocus>
<button type="submit">發送驗證碼</button></form>'''
    return f'<!DOCTYPE html><html lang="zh-HK"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>登入</title><style>{css}</style></head><body><div class="card">{body}</div></body></html>'


# ═══════════════ GOOGLE DRIVE LISTING ═══════════════

@app.route('/projects')
def projects_page():
    if not _check_auth():
        return '<script>location.href="/login"</script>'
    try:
        projects = list_projects()
        grouped = []
        seen_year = set()
        seen_month = set()
        for p in projects:
            date = p['date'] if p['date'] else 'unknown'
            year = date[:4]
            ym = date[:7]
            if year not in seen_year:
                seen_year.add(year)
                grouped.append({'type': 'year', 'label': year})
            if ym not in seen_month:
                seen_month.add(ym)
                grouped.append({'type': 'month', 'label': ym[5:7].lstrip('0') + '月',
                                'count': sum(1 for pp in projects if (pp.get('date', '') or '')[:7] == ym)})
            grouped.append({'type': 'item', 'data': p, 'ym': ym})
    except Exception as e:
        return f'<h1>無法連接 Google Drive</h1><p>{e}</p><a href="/">返回</a> <a href="https://drive.google.com" target="_blank" style="color:#3b82f6">開啟 Google Drive &#x2197;</a>', 500
    # Sync Drive files into local project tracker
    try:
        synced = sync_from_drive(projects)
        if synced > 0:
            print(f'Synced {synced} new projects from Drive')
    except Exception:
        pass
    return render_template('projects.html', grouped=grouped, version=_VERSION)


@app.route('/api/projects/list')
def api_project_list():
    if not _check_auth():
        return jsonify({'error': 'unauthorized'}), 401
    try:
        projects = list_projects()
        return jsonify(projects)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects/<file_id>')
def api_load_project(file_id):
    try:
        data = get_project_data(file_id)
        data['_filename'] = file_id
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════ LOCAL PROJECT MANAGEMENT ═══════════════

@app.route('/api/projects/local')
def api_local_projects():
    if not _check_auth():
        return jsonify({'error': 'unauthorized'}), 401
    status = request.args.get('status', '')
    projects = list_projects_local(status if status else None)
    return jsonify(projects)


@app.route('/api/projects/local/<project_id>')
def api_local_project(project_id):
    proj = get_project(project_id)
    if not proj:
        return jsonify({'error': 'not found'}), 404
    return jsonify(proj)


@app.route('/api/projects/local/<project_id>/status', methods=['POST'])
def api_update_status(project_id):
    body = request.get_json(silent=True) or {}
    new_status = body.get('status', '')
    if new_status not in ['draft', 'sent', 'confirmed', 'in_progress', 'completed', 'cancelled']:
        return jsonify({'error': 'invalid status'}), 400
    proj = update_status(project_id, new_status)
    if not proj:
        return jsonify({'error': 'not found'}), 404
    return jsonify(proj)


@app.route('/api/projects/local/<project_id>/payment/<int:idx>', methods=['POST'])
def api_toggle_payment(project_id, idx):
    proj = toggle_payment(project_id, idx)
    if not proj:
        return jsonify({'error': 'not found'}), 404
    return jsonify(proj)


# ═══════════════ DASHBOARD ═══════════════


@app.route('/api/sync-drive', methods=['POST'])
def api_sync_drive():
    """Parse Drive xlsx files into projects.json for Dashboard/專案管理"""
    if not _check_auth():
        return jsonify({'error': 'unauthorized'}), 401
    try:
        parsed = parse_drive_files(limit=20)
        return jsonify({'status': 'ok', 'parsed': parsed})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard')
def api_dashboard():
    if not _check_auth():
        return jsonify({'error': 'unauthorized'}), 401
    try:
        stats = get_dashboard_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════ HELPERS ═══════════════

def _make_filename(data, title):
    parts = []
    for key in ['quotation_no', 'address', 'owner_name']:
        v = (data.get(key, '') or '').strip()
        if v:
            parts.append(v)
    date = (data.get('date', '') or '').replace('-', '')
    if date.strip():
        parts.append(date.strip())
    if not parts:
        parts.append(date or '-')
    return '_'.join(parts) + '_' + title


def _parse_items_universal(ws):
    items = []
    col_desc, col_qty, col_price = _detect_columns(ws)
    skip = set(['項目', '編號', '描述', '總計', 'TOTAL', '小計', '合計', '訂金', '尾款', '備註',
                '地址', '日期', '名稱', '電話', '傳真', '公司', '客戶', 'No.', 'Item',
                'Qty', 'Unit', 'Price', 'Amount', 'Subtotal', '工程付款', '條款', '付款期數', '付款條件'])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if not row or row[0] is None:
            continue
        r = row[0].row
        a = str(ws.cell(row=r, column=1).value or '').strip()
        b = str(ws.cell(row=r, column=2).value or '').strip()
        if not a.isdigit() and 2 <= len(a) <= 6 and not any(k in a for k in skip):
            continue
        if b and not b[0].isdigit() and 2 <= len(b) <= 6 and not any(k in b for k in skip):
            continue
        desc = str(ws.cell(row=r, column=col_desc).value or '').strip()
        if (not desc or len(desc) < 2):
            a_text = str(ws.cell(row=r, column=1).value or '').strip()
            if len(a_text) > 3 and not a_text[0].isdigit():
                desc = a_text
        if not desc or len(desc) < 2:
            continue
        if any(k in desc for k in skip):
            continue
        seq = a
        is_seq = bool(re.match(r'^\d+[\.\)]?\s*$', seq) or re.match(r'^\d+\.\d+$', seq) or re.match(r'^\d+\)?$', seq))
        price_val = ws.cell(row=r, column=col_price).value
        has_price = isinstance(price_val, (int, float)) and float(price_val) >= 0
        is_item = (is_seq and len(desc) >= 2) or (has_price and len(desc) > 2)
        if is_item and not desc.startswith('='):
            try:
                qty = int(float(ws.cell(row=r, column=col_qty).value or 1))
            except Exception:
                qty = 1
            try:
                price = int(float(price_val)) if price_val else 0
            except Exception:
                price = 0
            items.append({'description': desc, 'quantity': qty, 'unit': '項',
                          'unit_price': price, 'remark': ''})
    return items


def _detect_columns(ws):
    cd, cq, cp = 2, 3, 5
    nc, tc = {}, {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
        r = row[0].row
        for c in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and float(v) > 0:
                nc[c] = nc.get(c, 0) + 1
            if isinstance(v, str) and len(v) > 3:
                tc[c] = tc.get(c, 0) + 1
    for c, _ in sorted(nc.items(), key=lambda x: -x[0]):
        if nc.get(c, 0) >= 2:
            cp = c
            break
    if tc:
        al = {}
        for c in tc:
            lens = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
                v = str(ws.cell(row=row[0].row, column=c).value or '')
                if len(v) > 1:
                    lens.append(len(v))
            al[c] = sum(lens) / max(len(lens), 1)
        for c, _ in sorted(tc.items(), key=lambda x: (-al.get(x[0], 0), -x[1])):
            if c < cp and c != 1 and tc.get(c, 0) >= 2 and al.get(c, 0) > 5:
                cd = c
                break
    for c in range(cd + 1, cp):
        if nc.get(c, 0) >= 2:
            cq = c
            break
    return cd, cq, cp


def _auto_categorize(desc):
    cats = {
        '清拆工程': ['清拆', '拆', '垃圾', '棚架', '保險'],
        '水電工程': ['水喉', '電制', '插蘇', '煤氣', '冷氣', '菲士', '供電', '燈'],
        '泥水工程': ['鋪磚', '磁磚', '瓷磚', '英泥沙', '防水', '地板', '地台', '企缸', '鋁窗', '窗台'],
        '油漆工程': ['油漆', '批灰', '油油', '起底', '鏟底', 'ICI', '乳膠漆'],
        '木工工程': ['門', '櫃', '床', '天花', '腳線', '廚櫃', '衣櫃', '雲石'],
        '安裝代工': ['安裝', '代裝', '人工'],
    }
    for cat, kws in cats.items():
        for kw in kws:
            if kw in desc:
                return cat
    return '雜項'


def _build_download_page(pid, title, data, fname, project_id=''):
    addr = data.get('address', '') or data.get('project_name', 'output')
    return f"""<!DOCTYPE html><html lang="zh-HK"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>{title}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#f0f2f5;font-family:'Microsoft JhengHei',sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;padding:20px}}
.card{{background:#fff;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.1);padding:36px;text-align:center;max-width:460px;width:100%}}
h1{{font-size:20px;color:#1e293b;margin-bottom:6px}}
.filename{{color:#64748b;font-size:13px;margin-bottom:20px;word-break:break-all}}
.btn{{display:block;width:100%;padding:13px;border:none;border-radius:8px;font-size:15px;cursor:pointer;font-weight:bold;color:#fff;margin-bottom:10px;text-decoration:none;text-align:center}}
.btn-excel{{background:#1F4E78}}
.btn-drive{{background:#0F9D58}}
.btn-whatsapp{{background:#25D366}}
.btn-back{{background:#94a3b8;font-size:13px;padding:10px}}
.btn-row{{display:flex;gap:8px;margin-bottom:10px}}
.btn-row .btn{{flex:1;font-size:14px}}
.ver{{color:#aaa;font-size:12px;margin-top:16px}}
.drive-msg{{color:#0F9D58;font-size:12px;margin-top:4px;display:none}}
.drive-err{{color:#ef4444;font-size:12px;margin-top:4px;display:none}}
</style></head><body>
<div class="card">
<h1>{title}已生成 &#x2705;</h1>
<p class="filename">{fname}.xlsx</p>
<a class="btn btn-excel" href="/download/{pid}/excel">&#x1F4E5; 下載 Excel</a>
<div class="btn-row">
<button class="btn btn-drive" onclick="saveDrive('{pid}')">&#x2601; 存去 Google Drive</button>
<button class="btn btn-whatsapp" onclick="shareWhatsApp('{pid}')">&#x1F4F1; WhatsApp 分享</button>
</div>
<div class="drive-msg" id="driveMsg"></div>
<div class="drive-err" id="driveErr"></div>
<a class="btn btn-back" href="/">返回主頁</a>
<a class="btn btn-back" href="/dashboard" style="margin-top:4px">&#x1F4CA; Dashboard</a>
<div class="ver">{_VERSION}</div>
</div>
<script>
async function saveDrive(pid){{
  const m=document.getElementById('driveMsg'),e=document.getElementById('driveErr');
  m.style.display='none';e.style.display='none';
  m.textContent='上傳中…';m.style.display='block';
  try{{
    const r=await fetch('/save-drive/'+pid,{{method:'POST'}});
    if(!r.ok){{const j=await r.json();throw new Error(j.error||'failed')}}
    const j=await r.json();
    m.innerHTML='已存到 Drive！<a href="'+j.web_link+'" target="_blank">&#x1F517; 開啟</a>';
  }}catch(err){{e.textContent='失敗：'+err.message;e.style.display='block';m.style.display='none'}}
}}
async function shareWhatsApp(pid){{
  try{{
    const r=await fetch('/api/whatsapp/'+pid);
    if(!r.ok)return alert('無法獲取分享資料');
    const j=await r.json();
    window.open('https://wa.me/?text='+encodeURIComponent(j.text),'_blank');
  }}catch(err){{alert('分享失敗：'+err.message)}}
}}
</script>
</body></html>"""


# ═══════════════ MAIN ═══════════════

if __name__ == '__main__':
    print('=' * 50)
    print(f'裝修報價單/發票助手 {_VERSION}')
    print('http://localhost:5000')
    print('=' * 50)
    app.run(debug=True, host='127.0.0.1', port=5000)
