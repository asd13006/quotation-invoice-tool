"""Google Drive sync — list + upload xlsx files"""
import io
import os
import json as _json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook

SCOPES = ['https://www.googleapis.com/auth/drive.file']
_drive = None


def _get_drive():
    global _drive
    if _drive is None:
        sa_json = os.environ.get('SERVICE_ACCOUNT_JSON', '')
        if sa_json:
            info = _json.loads(sa_json)
            creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
        else:
            creds = service_account.Credentials.from_service_account_file(
                'service-account.json', scopes=SCOPES)
        _drive = build('drive', 'v3', credentials=creds)
    return _drive


def list_projects():
    drive = _get_drive()
    results = drive.files().list(
        q="mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        pageSize=100,
        fields='files(id, name, parents, modifiedTime, size)',
        orderBy='modifiedTime desc'
    ).execute()
    projects = []
    for f in results.get('files', []):
        projects.append({
            'id': f['id'],
            'name': f['name'].replace('.xlsx', ''),
            'date': (f.get('modifiedTime', '') or '')[:10],
            'size_kb': int(f.get('size', 0)) // 1024,
            'folder_id': (f.get('parents', ['']))[0],
        })
    return projects


def get_project_data(file_id):
    drive = _get_drive()
    request = drive.files().get_media(fileId=file_id)
    xlsx_bytes = request.execute()
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    data = {}
    header_map = {
        '工程名稱': 'project_name', '客戶姓名': 'owner_name',
        '工程地址': 'address', '裝修公司': 'company_name',
        '報價單號': 'quotation_no', '報價日期': 'date',
        '有效期': 'validity', '版本': 'version',
    }
    for r in range(1, 11):
        for c in range(1, 8):
            label = str(ws.cell(row=r, column=c).value or '')
            for kw, key in header_map.items():
                if kw in label:
                    val = str(ws.cell(row=r, column=c + 1).value or '')
                    if val and val not in ('None', '-', ''):
                        data[key] = val
    data['items'] = []
    import re
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        if not row or row[0] is None:
            continue
        r = row[0].row
        a = str(ws.cell(row=r, column=1).value or '').strip()
        b = str(ws.cell(row=r, column=2).value or '').strip()
        is_seq = bool(re.match(r'^\d+[\.\)]?\s*$', a) or re.match(r'^\d+\.\d+$', a))
        if is_seq and len(b) > 2:
            e = ws.cell(row=r, column=5).value
            price = int(e) if isinstance(e, (int, float)) else 0
            c_val = ws.cell(row=r, column=3).value
            qty = int(c_val) if isinstance(c_val, (int, float)) and float(c_val) > 0 else 1
            data['items'].append({
                'category': '雜項', 'description': b,
                'quantity': qty, 'unit': '項', 'unit_price': price,
                'remark': '', 'is_additional': False,
            })
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and '訂金' in str(cell.value):
                fv = ws.cell(row=cell.row, column=6).value
                if isinstance(fv, (int, float)):
                    data['deposit'] = int(fv)
    data['payments'] = []
    data['terms'] = []
    return data


def upload_file(xlsx_bytes, filename, folder_id=None):
    """Upload xlsx to Google Drive, return file metadata"""
    drive = _get_drive()
    file_meta = {'name': filename, 'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
    if folder_id:
        file_meta['parents'] = [folder_id]
    media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=file_meta['mimeType'], resumable=True)
    f = drive.files().create(body=file_meta, media_body=media, fields='id,name,webViewLink').execute()
    return {'id': f.get('id', ''), 'name': f.get('name', ''), 'web_link': f.get('webViewLink', '')}
